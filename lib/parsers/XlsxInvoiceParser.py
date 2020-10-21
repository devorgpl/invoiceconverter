import re

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from lib.InvoiceDocument import InvoiceDocument, InvoiceHeader, InvoiceParty, InvoiceSummary, InvoiceTaxSummary, \
    InvoiceLine
from lib.ParserInterface import ParserInterface


class XlsxInvoiceParser(ParserInterface):

    def test(self, filename: str) -> bool:
        return filename.endswith(".xlsx")

    def parse(self, filename: str) -> InvoiceDocument:
        wb = load_workbook(filename)
        result = InvoiceDocument()
        result.header = self.parseHeader(wb['Header'])
        result.buyer = self.parseParty(wb['Buyer'])
        result.seller = self.parseParty(wb['Seller'])
        result.summary = self.parseSummary(wb['Summary'])
        result.lines = self.parseLines(wb['Order'])
        # result = self.processResult(result)
        return result

    def parseLines(self, sheet: Worksheet):
        result = []
        rows = sheet.iter_rows()
        next(rows)
        for row in rows:
            if row[0].value is not None:
                result.append(self.parseLine(row))
        return result

    def parseHeader(self, sheet: Worksheet) -> InvoiceHeader:
        header = InvoiceHeader()
        header.number = sheet['A2'].value
        header.invoiceDate = sheet['B2'].value
        header.salesDate = sheet['B2'].value
        header.currency = sheet['D2'].value
        header.paymentDueDate = sheet['B2'].value
        header.paymentTerms = ""
        header.documentFunctionCode = ""
        return header

    def parseParty(self, sheet: Worksheet) -> InvoiceParty:
        party = InvoiceParty()
        party.taxId = str(sheet['A2'].value)
        party.accountNumber = sheet['B2'].value
        party.name = sheet['C2'].value
        party.streetAndNumber = sheet['D2'].value
        party.cityName = sheet['E2'].value
        party.postalCode = sheet['F2'].value
        party.country = sheet['G2'].value
        return party

    def parseSummary(self, sheet: Worksheet) -> InvoiceSummary:
        summary = InvoiceSummary()
        summary.totalLines = str(sheet['A2'].value)
        summary.totalNetAmount = str(sheet['B2'].value)
        summary.totalTaxableBasis = ""
        summary.totalTaxAmount = str(sheet['C2'].value)
        summary.totalGrossAmount = str(sheet['D2'].value)
        summary.grossAmountInWords = ""
        # summary.taxSummary = self.parseTaxSummary(xmlSummary.getElementsByTagName("Tax-Summary-Line")[0])
        return summary

    def parseTaxSummary(self, sheet: Worksheet) -> InvoiceTaxSummary:
        tax_summary = InvoiceTaxSummary()
        # tax_summary.taxRate = xmlExtractText(xmlTaxSummary, 'TaxRate')
        # tax_summary.taxCategoryCode = xmlExtractText(xmlTaxSummary, 'TaxCategoryCode')
        # tax_summary.taxAmount = xmlExtractText(xmlTaxSummary, 'TaxAmount')
        # tax_summary.taxableBasis = xmlExtractText(xmlTaxSummary, 'TaxableBasis')
        # tax_summary.taxableAmount = xmlExtractText(xmlTaxSummary, 'TaxableAmount')
        # tax_summary.grossAmount = xmlExtractText(xmlTaxSummary, 'GrossAmount')
        return tax_summary

    def parseLine(self, row) -> InvoiceLine:
        line = InvoiceLine()
        line.LineNumber = str(row[0].value)
        line.EAN = str(row[2].value)
        line.BuyerItemCode = str(row[2].value)
        line.SupplierItemCode = str(row[2].value)
        line.ItemDescription = str(row[1].value)
        line.ItemType = ""
        line.InvoiceQuantity = str(row[3].value)
        line.UnitOfMeasure = "szt."
        line.InvoiceUnitPacksize = str(1)
        line.PackItemUnitOfMeasure = "szt."
        line.InvoiceUnitNetPrice = str(round(row[10].value, 2))
        line.TaxRate = str(row[7].value)
        line.TaxCategoryCode = ""
        line.ReferenceType = ""
        line.ReferenceNumber = ""
        line.TaxAmount = str(round(row[9].value * row[7].value / 100, 2))
        line.NetAmount = str(row[9].value)
        return line

    def processResult(self, result):
        result.buyer.shortName = self.shortName(result.buyer.name)
        result.seller.shortName = self.shortName(result.seller.name)
        return result

    def shortName(self, name):
        match = re.search('\\A[@\\w]+', name)
        if match:
            return match.group()
        else:
            return name
